import winreg
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import shutil
import pandas as pd
import re

def get_display_name(oKey, debug=False):
    candidates = [
        "DisplayName",
        "ProductName",
        ("InstallProperties", "DisplayName")
    ]
    
    for candidate in candidates:
        try:
            if isinstance(candidate, tuple):
                subkey_name, value_name = candidate
                subKey = winreg.OpenKey(oKey, subkey_name)
                value, _ = winreg.QueryValueEx(subKey, value_name)
            else:
                value, _ = winreg.QueryValueEx(oKey, candidate)
            return value
        except EnvironmentError as err:
            if debug:
                print(f"Failed reading {candidate}: {err}")
            continue
    return "None"

def findDisplayNames(path, number_of_keys=4096, user=False, debug=False):
    hive = winreg.HKEY_CURRENT_USER if user else winreg.HKEY_LOCAL_MACHINE
    aReg = winreg.ConnectRegistry(None, hive)
    aKey = winreg.OpenKey(aReg, path)
    
    print(r"*** Reading from %s ***" % path)
    returnArray = []
    
    for i in range(number_of_keys):
        try:
            folder_name = winreg.EnumKey(aKey, i)
            oKey = winreg.OpenKey(aKey, folder_name)
            display_name = get_display_name(oKey, debug)
            returnArray.append(f"{folder_name} \\ {display_name}")
            if debug:
                print(f'Appending "{folder_name} \\ {display_name}"')
        except EnvironmentError as err:
            if debug:
                print(err)
            if "No more data is available" in str(err):
                break
            elif "The system cannot find the file specified" in str(err):
                continue
            print(err)
    return returnArray

def save_excel(path, file_name, data, screen_column_count = 29, sheet_name = "data"):
    default_column_width = 8.43
    total_available_width = screen_column_count * default_column_width
    
    if os.path.exists(f"{path}\\{file_name}.xlsx"):
        print("File exists")
        wb = load_workbook(f"{path}\\{file_name}.xlsx")
        if sheet_name in wb.sheetnames:
            print("Sheet exists, removing old version")
            ws_to_remove = wb[sheet_name]
            wb.remove(ws_to_remove)
            ws = wb.create_sheet(sheet_name)
        else:
            print("Sheet does not exist, creating new sheet")
            ws = wb.create_sheet(sheet_name)
    else:
        print("File doesn't exist, creating new workbook")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    print("Writing data")
    ws.cell(1, 1, file_name)

    for col_index, col in enumerate(data, start=2):
        for row_index, value in enumerate(col, start=1):
            ws.cell(row_index, col_index, value)

    if data:
        max_cols = len(data) + 1
    else:
        max_cols = 0

    if max_cols > 0:
        new_width = total_available_width / max_cols
        for col in range(1, max_cols + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = new_width

    print("Saving file")
    wb.save(f"{path}\\{file_name}.xlsx")

def move_old_excel_files(path, today_str):
    today_date = datetime.strptime(today_str, "%Y-%m-%d").date()
    
    excel_files = [f for f in os.listdir(path) if os.path.isfile(f) and f.lower().endswith(('.xlsx', '.xls'))]

    keep_files = set()
    date_files = []
    
    for file in excel_files:
        base, ext = os.path.splitext(file)
        
        if base == today_str:
            keep_files.add(file)
        else:
            try:
                file_date = datetime.strptime(base, "%Y-%m-%d").date()

                if file_date < today_date:
                    date_files.append((file_date, file))
            except ValueError:
                pass

    if date_files:
        latest_date, latest_file = max(date_files, key=lambda x: x[0])
        keep_files.add(latest_file)
    
    old_folder = os.path.join(path, "old")
    if not os.path.exists(old_folder):
        os.makedirs(old_folder)
    
    for file in excel_files:
        if file not in keep_files:
            src = os.path.join(path, file)
            dst = os.path.join(old_folder, file)
            print(f"Moving {file} to {dst}")
            shutil.move(src, dst)

def read_and_compare(folder_path: str, current_date: str):
    excel_files = [f for f in os.listdir(folder_path) if re.match(r'\d{4}-\d{2}-\d{2}\.xlsx$', f)]
    
    if len(excel_files) != 2:
        raise ValueError("Folder must contain exactly two Excel files with the format Y-M-D.xlsx")

    excel_files.sort()
    old_file, new_file = excel_files if current_date in excel_files[1] else excel_files[::-1]

    old_df = pd.read_excel(os.path.join(folder_path, old_file), dtype=str)
    new_df = pd.read_excel(os.path.join(folder_path, new_file), dtype=str)

    print(f"\n=========== Comparing {old_file} to {new_file} ===========\n")

    old_columns = set(old_df.columns)
    new_columns = set(new_df.columns)

    added_columns = new_columns - old_columns
    removed_columns = old_columns - new_columns

    for col in removed_columns:
        print(f"- Column removed: {col}")

    for col in added_columns:
        print(f"+ Column added: {col}")

    common_columns = old_columns & new_columns

    for col in common_columns:
        old_values = set(old_df[col].dropna())
        new_values = set(new_df[col].dropna())

        removed_values = old_values - new_values
        added_values = new_values - old_values

        if removed_values or added_values:
            print(f"\nChanges in column: {col}")

        for value in removed_values:
            print(f"- {col}: {value}")

        for value in added_values:
            print(f"+ {col}: {value}")

def main():
    file_path = "C:\\Users\\Swoots\\Desktop\\Reg-Track"
    today = datetime.now().strftime("%Y-%m-%d")

    regestry_folders = [
        ("SOFTWARE\\WOW6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall", True),
        ("SOFTWARE\\WOW6432Node\\Microsoft\\Windows\\CurrentVersion\\Uninstall", False),
        ("SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall", True),
        ("SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Uninstall", False),
        ("SOFTWARE\\WOW6432Node", True),
        ("SOFTWARE\\WOW6432Node", False),
        ("SOFTWARE", True),
        ("SOFTWARE", False),
        ("SOFTWARE\\Classes\\Installer\\Products", False),
        ("SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Installer\\UserData\\S-1-5-18\\Products", False),
    ]

    sheet_data = []

    for regestry_folder in regestry_folders:
        data = findDisplayNames(regestry_folder[0])

        if regestry_folder[1]:
            data.insert(0, f"HKEY_CURRENT_USER\\{regestry_folder[0]}")
        else:
            data.insert(0, f"HKEY_LOCAL_MACHINE\\{regestry_folder[0]}")

        sheet_data.append(data)

    save_excel(file_path, today, sheet_data)
    move_old_excel_files(file_path, today)
    read_and_compare(file_path, today)

    input('\nPress Enter to exit: ')

if __name__ == "__main__":
    main()